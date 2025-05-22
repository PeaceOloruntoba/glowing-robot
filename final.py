import tweepy
import pandas as pd
import nltk
from nltk.sentiment import SentimentIntensityAnalyzer
from nltk.tokenize import word_tokenize
from textblob import TextBlob
import re
import json
from datetime import datetime
import pymongo
import redis
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# Download required NLTK data
nltk.download('vader_lexicon')
nltk.download('punkt')

class UniversitySentimentAnalyzer:
    def __init__(self):
        # Twitter API credentials
        self.consumer_key = 'yFhdFvUmrYKYWS58BpaUTQW8H'
        self.consumer_secret = 'your_consumer_secret_here'
        self.access_token = '1615778868824096768-SceLOd38jaBbvGNsNRC2xjSL8uRkiu'
        self.access_token_secret = 'ojoPCjU0HCs4kxOvk2bRKHLC8gJiZQLZ1QZg2ZdSY7FN7'
        
        # Initialize Twitter API
        auth = tweepy.OAuthHandler(self.consumer_key, self.consumer_secret)
        auth.set_access_token(self.access_token, self.access_token_secret)
        self.api = tweepy.API(auth)
        
        # Initialize sentiment analyzer
        self.sia = SentimentIntensityAnalyzer()
        
        # Connect to MongoDB
        self.mongo_client = pymongo.MongoClient("mongodb://localhost:27017/")
        self.db = self.mongo_client["university_sentiments"]
        
        # Connect to Redis
        self.redis_client = redis.Redis(host='localhost', port=6379, db=0)

    def preprocess_text(self, text):
        """Clean and preprocess tweet text"""
        text = re.sub(r'https?://\S+', '', text)  # Remove URLs
        text = re.sub(r'@\w+', '', text)  # Remove mentions
        text = re.sub(r'#', '', text)  # Remove hashtags
        text = re.sub(r'\n', ' ', text)  # Replace newlines with spaces
        return text.strip()

    def fetch_tweets(self, query, count=100):
        """Fetch tweets for a specific query"""
        tweets = []
        try:
            for tweet in tweepy.Cursor(
                self.api.search_tweets,
                q=query,
                lang='en',
                result_type='recent'
            ).items(count):
                processed_tweet = {
                    'text': self.preprocess_text(tweet.text),
                    'created_at': tweet.created_at,
                    'username': tweet.user.screen_name,
                    'location': tweet.user.location,
                    'source': 'twitter'
                }
                tweets.append(processed_tweet)
        except Exception as e:
            print(f"Error fetching tweets: {str(e)}")
        
        return tweets

    def analyze_sentiment(self, text):
        """Analyze sentiment using multiple methods"""
        # NLTK VADER
        nltk_scores = self.sia.polarity_scores(text)
        
        # TextBlob
        blob = TextBlob(text)
        polarity = blob.sentiment.polarity
        subjectivity = blob.sentiment.subjectivity
        
        return {
            'nltk_compound': nltk_scores['compound'],
            'nltk_pos': nltk_scores['pos'],
            'nltk_neg': nltk_scores['neg'],
            'nltk_neutral': nltk_scores['neu'],
            'textblob_polarity': polarity,
            'textblob_subjectivity': subjectivity
        }

    def store_data(self, university, tweets):
        """Store processed tweets in MongoDB"""
        collection = self.db[university]
        
        for tweet in tweets:
            sentiment_scores = self.analyze_sentiment(tweet['text'])
            tweet.update(sentiment_scores)
            
            # Store in MongoDB
            collection.insert_one(tweet)
            
            # Cache recent sentiment scores in Redis
            self.redis_client.hset(
                f"{university}:sentiments",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                json.dumps(sentiment_scores)
            )

    def generate_visualizations(self):
        """Generate visualizations for analysis"""
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        
        for idx, university in enumerate(['Anchor University', 'University of Lagos']):
            collection = self.db[university.replace(" ", "_").lower()]
            tweets = list(collection.find())
            
            # Convert to DataFrame
            df = pd.DataFrame(tweets)
            
            # Sentiment distribution plot
            sns.histplot(data=df, x='nltk_compound', ax=axes[0, idx], bins=20)
            
            # Time series sentiment plot
            df['created_at'] = pd.to_datetime(df['created_at'])
            df.set_index('created_at').resample('D')['nltk_compound'].mean().plot(
                ax=axes[1, idx], marker='o'
            )
            
        plt.tight_layout()
        return fig

    def export_to_excel(self):
        """Export analysis results to Excel"""
        wb = Workbook()
        ws = wb.active
        
        # Create header row
        headers = ['University', 'Tweet Text', 'Sentiment Score', 'Date']
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)
        
        row = 2
        for university in ['Anchor University', 'University of Lagos']:
            collection = self.db[university.replace(" ", "_").lower()]
            tweets = list(collection.find())
            
            for tweet in tweets:
                ws.cell(row=row, column=1, value=university)
                ws.cell(row=row, column=2, value=tweet['text'])
                ws.cell(row=row, column=3, value=tweet['nltk_compound'])
                ws.cell(row=row, column=4, value=tweet['created_at'].strftime('%Y-%m-%d %H:%M:%S'))
                row += 1
        
        # Add charts
        chart_ws = wb.create_sheet('Charts')
        
        # Create bar chart for average sentiments
        anchor_avg = self.db['anchor_university'].aggregate([
            {'$group': {'_id': None, 'avg_sentiment': {'$avg': '$nltk_compound'}}}
        ])[0]['avg_sentiment']
        
        unilag_avg = self.db['university_of_lagos'].aggregate([
            {'$group': {'_id': None, 'avg_sentiment': {'$avg': '$nltk_compound'}}}
        ])[0]['avg_sentiment']
        
        data = Reference(chart_ws, min_col=1, min_row=1, max_row=2)
        chart = BarChart()
        chart.add_data(data)
        chart_ws.add_chart(chart, "E2")
        
        wb.save('university_sentiment_analysis.xlsx')
        
        
        
# Initialize analyzer
analyzer = UniversitySentimentAnalyzer()

# Define search queries
queries = [
    '"Anchor University Lagos infrastructure" OR "Anchor University facilities"',
    '"University of Lagos infrastructure" OR "UNILAG facilities"'
]

# Fetch and analyze tweets for both universities
for university, query in zip(['Anchor University', 'University of Lagos'], queries):
    print(f"\nFetching tweets for {university}...")
    tweets = analyzer.fetch_tweets(query=query, count=100)
    analyzer.store_data(university, tweets)

# Generate visualizations and export results
fig = analyzer.generate_visualizations()
plt.savefig('sentiment_analysis_plots.png')
analyzer.export_to_excel()